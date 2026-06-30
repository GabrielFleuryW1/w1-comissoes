"""
W1 Consultoria – Controle Unificado de Comissões + Bônus
Streamlit App: consultor faz upload do CSV e recebe planilha completa.

Autor: Gabriel Fleury / Claude
"""
import streamlit as st
import pandas as pd
import numpy as np
import re, io, warnings, requests, base64, json
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════════════════════════════════════
# USAGE LOGGING (GitHub CSV)
# ═══════════════════════════════════════════════════════════════════════════════
def log_usage(nome, cargo, n_contratos, total_hist, total_proj):
    """Registra cada uso no arquivo usage_log.csv do repositório GitHub."""
    try:
        token = st.secrets.get("GITHUB_TOKEN", "")
        if not token:
            return
        owner, repo, path = "GabrielFleuryW1", "w1-comissoes", "usage_log.csv"
        api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"
        headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        new_line = f"{now},{nome},{cargo},{n_contratos},{total_hist:.2f},{total_proj:.2f}\n"
        resp = requests.get(api_url, headers=headers, timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            existing = base64.b64decode(data["content"]).decode("utf-8")
            sha = data["sha"]
            content = existing + new_line
        else:
            sha = None
            content = "data_hora,nome,cargo,contratos,total_historico,total_projetado\n" + new_line
        payload = {"message": f"log: {nome} {now}", "content": base64.b64encode(content.encode()).decode(), "branch": "main"}
        if sha:
            payload["sha"] = sha
        requests.put(api_url, headers=headers, json=payload, timeout=10)
    except Exception:
        pass  # silencioso — nunca interrompe o app

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS & STYLES
# ═══════════════════════════════════════════════════════════════════════════════
C_N="1B3A5C"; C_B="2E6DA4"; C_G="1A7A4A"; C_R="C0392B"; C_P="6B3FA0"
C_GL="C9A84C"; C_R1="EBF2FA"; C_R2="FFFFFF"; C_LG="E8F5E9"
FN="Arial"; FB='R$ #,##0.00'
FA_RATES = {'FA I': 0.075, 'FA II': 0.15, 'FA III': 0.175, 'FA IV': 0.20}
BONUS_TOTAL = {'2000': 0.40, '3000': 0.45}

# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════
def hdr(ws,r,c,v,bg=C_N,fg="FFFFFF",sz=9,bold=True,wrap=True,al="center"):
    cl=ws.cell(row=r,column=c,value=v)
    cl.font=Font(name=FN,bold=bold,color=fg,size=sz)
    cl.fill=PatternFill("solid",fgColor=bg)
    cl.alignment=Alignment(horizontal=al,vertical="center",wrap_text=wrap)
    return cl

def dc(ws,r,c,v,fmt=None,bold=False,bg=None,fg="000000",al="left",wrap=False,sz=8):
    cl=ws.cell(row=r,column=c,value=v)
    cl.font=Font(name=FN,size=sz,bold=bold,color=fg)
    cl.alignment=Alignment(horizontal=al,vertical="center",wrap_text=wrap)
    if bg: cl.fill=PatternFill("solid",fgColor=bg)
    if fmt: cl.number_format=fmt
    return cl

def thin():
    s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)

def borders(ws,r1,r2,c1,c2):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            ws.cell(r,c).border=thin()

def set_cols(ws,w):
    for col,width in w.items(): ws.column_dimensions[col].width=width

def norm_prod(p):
    if pd.isna(p): return ''
    return str(p).strip().replace('Imóveis','Imóvel').replace('Automóveis','Automóvel')

def norm_name(n):
    if pd.isna(n): return ''
    return str(n).strip().lower()

# ═══════════════════════════════════════════════════════════════════════════════
# CORE ENGINE: build the unified spreadsheet
# ═══════════════════════════════════════════════════════════════════════════════
def build_unified(csv_bytes, params):
    """
    Main engine. Takes CSV bytes and a params dict with:
      - nome: str (full name as in CSV)
      - promos: dict of {'FA II': 'YYYY-MM', 'FA III': 'YYYY-MM', 'FA IV': 'YYYY-MM'} or None
      - first_bonus_2000: 'YYYY-MM' or None (first BASE month that generates Bônus 2000)
      - first_bonus_3000: 'YYYY-MM' or None (first BASE month that generates Bônus 3000)
      - holding_parcelas: int or 0
      - holding_valor: float or 0
      - holding_pagas: int or 0
      - business_valor: float or 0 (0 = no business)
      - exclude_ap_bonus: bool (exclude Bônus Cliente de Alto Padrão)
    Returns: (xlsx_bytes, summary_dict)
    """
    NM = params['nome']

    # Detect encoding (W1 exports are cp1252/latin-1)
    for enc in ['utf-8-sig', 'cp1252', 'latin-1']:
        try:
            df = pd.read_csv(io.BytesIO(csv_bytes), sep=';', encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        df = pd.read_csv(io.BytesIO(csv_bytes), sep=';', encoding='latin-1', encoding_errors='replace')

    # ── Extract value & FA position for this person ──
    def gvp(row):
        for cn,cv,pos in [('Nome FA I','Valor FA I','FA I'),('Nome FA II','Valor FA II','FA II'),
                           ('Nome FA III','Valor FA III','FA III'),('Nome FA IV','Valor FA IV','FA IV')]:
            if cn in row.index and row[cn]==NM:
                try: return float(str(row[cv]).replace(',','.')),pos
                except: return 0.0,pos
        return 0.0,None

    res=df.apply(gvp,axis=1,result_type='expand')
    df['VG']=res[0]; df['FA']=res[1]
    df['DT']=pd.to_datetime(df['Data da Movimentacão'],errors='coerce')
    df['YM']=df['DT'].dt.to_period('M')
    df['DI']=pd.to_datetime(df['Data Inicial'],errors='coerce')
    def pp(p):
        try: return int(str(p))
        except: return 0
    df['PN']=df['Parcela'].apply(pp)
    ent=df[(df['Tipo da Movimentação']=='Entrada')&(df['VG']>0)].copy()

    if len(ent)==0:
        return None, {'error': f'Nenhuma entrada encontrada para "{NM}". Verifique se o nome está exatamente igual ao do extrato.'}

    # ── Determine date range from data ──
    min_ym = str(ent['YM'].min())
    max_ym = str(ent['YM'].max())

    # Build month list: from min_ym to +9 months after max
    def ym_range(start, end):
        s = pd.Period(start, 'M')
        e = pd.Period(end, 'M')
        months = []
        while s <= e:
            months.append(str(s))
            s += 1
        return months

    proj_end_ym = str(pd.Period(max_ym,'M') + 9)
    ALL_YM = ym_range(min_ym, proj_end_ym)

    # Labels
    MONTH_NAMES = {1:'Jan',2:'Fev',3:'Mar',4:'Abr',5:'Mai',6:'Jun',
                   7:'Jul',8:'Ago',9:'Set',10:'Out',11:'Nov',12:'Dez'}
    ALL_LBL = []
    for ym in ALL_YM:
        p = pd.Period(ym,'M')
        ALL_LBL.append(f"{MONTH_NAMES[p.month]}/{str(p.year)[2:]}")

    N_HIST = ALL_YM.index(max_ym) + 1
    HIST_YM = ALL_YM[:N_HIST]
    PROJ_YM = ALL_YM[N_HIST:]

    # Determine first/last bonus base month
    FIRST_BONUS_BASE = params.get('first_bonus_2000') or params.get('first_bonus_3000')
    # Last auditable: one month before last historical month (because bonus = M+1)
    if N_HIST >= 2:
        LAST_BONUS_BASE = ALL_YM[N_HIST - 2]
    else:
        LAST_BONUS_BASE = max_ym

    # ── Parse bonus entries ──
    bonus_all = ent[ent['Categoria'].str.contains('Bônus', na=False)].copy()

    def parse_obs(obs):
        if pd.isna(obs) or not obs: return None,None,None,None
        obs=str(obs)
        m=re.match(r'^(.+?)\s*-\s*(\d{4}-\d{2}-\d{2})\s*-\s*(.+?)\s*-\s*(Bônus\s+\d+)$', obs)
        if m: return m.group(1).strip(), m.group(2), m.group(3).strip(), m.group(4).strip()
        return None,None,None,None

    if bonus_all.empty:
        bonus_all = bonus_all.assign(B_Prod=None, B_Data=None, B_Cliente=None, B_Tipo=None)
    else:
        bonus_all[['B_Prod','B_Data','B_Cliente','B_Tipo']] = bonus_all['Observações'].apply(
            lambda x: pd.Series(parse_obs(x)))
    bonus_all['B_Prod_N']=bonus_all['B_Prod'].apply(norm_prod)
    bonus_all['B_Cliente_N']=bonus_all['B_Cliente'].apply(norm_name)
    bonus_all['B_Ref_YM']=bonus_all['B_Data'].apply(
        lambda x: str(pd.to_datetime(x).to_period('M')) if pd.notna(x) and x else None)

    b2000=bonus_all[bonus_all['Categoria']=='Bônus 2000']
    b3000=bonus_all[bonus_all['Categoria']=='Bônus 3000']
    b_ap=bonus_all[bonus_all['Categoria']=='Bônus Cliente de Alto Padrão']

    bonus_lookup={}
    for bdf,tipo in [(b2000,'2000'),(b3000,'3000')]:
        for _,r in bdf.iterrows():
            ym=str(r['YM'])
            key=(ym, r['B_Cliente_N'], r['B_Prod_N'])
            if key not in bonus_lookup: bonus_lookup[key]=[]
            bonus_lookup[key].append({'valor':r['VG'],'tipo':tipo,'used':False})

    # ── Base products (exclude bonus entries) ──
    base_all=ent[ent['Classificação']=='Produtos Parceiros'].copy()
    base_all=base_all[~base_all['Categoria'].str.contains('Bônus',na=False)]
    excl_pat='Seguro Automóvel|Abertura de Conta|Educação Financeira'
    base_all=base_all[~base_all['Produto / Serviço'].str.contains(excl_pat,na=False)]

    # ═══════════════════════════════════════════════════════════════════════════
    # BUILD CONTRACT LIST
    # ═══════════════════════════════════════════════════════════════════════════
    contracts=[]

    # ── CONSÓRCIO ──
    cons=base_all[base_all['Produto / Serviço'].str.contains('Consórcio|Klubi',na=False)].copy()
    for (contato,prod),grp in cons.groupby(['Contato','Produto / Serviço']):
        fa=grp['FA'].mode().iloc[0] if len(grp)>0 else 'FA III'
        parc_counts=grp.groupby('PN')['VG'].count()
        n_cartas=int(parc_counts.max()) if len(parc_counts)>0 else 1
        per_carta_val=grp['VG'].iloc[0] if len(grp)>0 else 0
        monthly=grp.groupby(grp['YM'].astype(str))['VG'].sum().to_dict()
        max_parc=grp['PN'].max()
        data_ini=grp['DI'].dropna().min() if grp['DI'].notna().any() else None
        for carta_i in range(n_cartas):
            carta_monthly={ym:round(v/n_cartas,2) for ym,v in monthly.items()}
            contracts.append({
                'tipo':'Consórcio','contato':contato,'produto':prod,
                'detalhe':f'Carta {carta_i+1}/{n_cartas}','fa':fa,
                'val_mensal':per_carta_val,'total_parc':12,
                'parc_atual':max_parc,'data_ini':data_ini,
                'monthly':carta_monthly,'bonus_elegivel':True
            })

    # ── MAG VIDA ──
    mag=base_all[base_all['Produto / Serviço'].str.contains('MAG|Vida Total',na=False)].copy()
    for contato,grp in mag.groupby('Contato'):
        fa=grp['FA'].mode().iloc[0] if len(grp)>0 else 'FA III'
        monthly=grp.groupby(grp['YM'].astype(str))['VG'].sum().to_dict()
        max_parc=grp['PN'].max()
        n_cob=grp.groupby('PN')['VG'].count().max()
        recent_vals=[v for ym,v in sorted(monthly.items()) if ym>=max_ym[:5]]
        val_y1=np.mean(recent_vals) if recent_vals else np.mean(list(monthly.values()))
        data_ini=grp['DI'].dropna().min() if grp['DI'].notna().any() else None
        contracts.append({
            'tipo':'MAG Vida','contato':contato,'produto':'MAG Vida Private',
            'detalhe':f'{int(n_cob)} cob.','fa':fa,
            'val_mensal':val_y1,'total_parc':120,'parc_atual':max_parc,
            'data_ini':data_ini,'monthly':monthly,'bonus_elegivel':True,
            'val_y2':val_y1*0.36
        })

    # ── HORIZONTE / PREV ──
    horiz=base_all[base_all['Produto / Serviço'].str.contains('Horizonte|Zurich Prev|Zurich Port|Prev XP|Zurich Aporte',na=False)].copy()
    for (contato,prod),grp in horiz.groupby(['Contato','Produto / Serviço']):
        fa=grp['FA'].mode().iloc[0] if len(grp)>0 else 'FA III'
        monthly=grp.groupby(grp['YM'].astype(str))['VG'].sum().to_dict()
        recent_vals=[v for ym,v in sorted(monthly.items()) if ym>=ALL_YM[max(0,N_HIST-3)]]
        val_med=np.mean(recent_vals) if recent_vals else np.mean(list(monthly.values()))
        data_ini=grp['DI'].dropna().min() if grp['DI'].notna().any() else None
        contracts.append({
            'tipo':'Horizonte/Prev','contato':contato,'produto':prod,
            'detalhe':'Recorrente','fa':fa,
            'val_mensal':val_med,'total_parc':999,
            'parc_atual':grp['PN'].max(),'data_ini':data_ini,
            'monthly':monthly,'bonus_elegivel':True
        })

    # ── AP ──
    ap=ent[ent['Categoria']=='AP'].copy()
    for contato,grp in ap.groupby('Contato'):
        fa=grp['FA'].mode().iloc[0] if len(grp)>0 else 'FA IV'
        monthly=grp.groupby(grp['YM'].astype(str))['VG'].sum().to_dict()
        max_parc=grp['PN'].max()
        val_med=grp['VG'].mean()
        data_ini=grp['DI'].dropna().min() if grp['DI'].notna().any() else None
        ult_dt=grp['DT'].max()
        if grp['PN'].max()<=1 and len(grp)<=2 and val_med>500:
            detalhe='À vista (concluído)'
        else:
            detalhe=f'Parc {max_parc}/12'
        contracts.append({
            'tipo':'AP','contato':contato,'produto':'Assessoria Patrimonial',
            'detalhe':detalhe,'fa':fa,'val_mensal':val_med,'total_parc':12,
            'parc_atual':max_parc,'data_ini':data_ini,'monthly':monthly,
            'bonus_elegivel':False,'ult_dt':ult_dt
        })

    # ── FUP ──
    fup=ent[ent['Categoria']=='Acompanhamento'].copy()
    for contato,grp in fup.groupby('Contato'):
        fa=grp['FA'].mode().iloc[0] if len(grp)>0 else 'FA IV'
        monthly=grp.groupby(grp['YM'].astype(str))['VG'].sum().to_dict()
        val_med=grp['VG'].mean()
        ult_dt=grp['DT'].max()
        cutoff = pd.Timestamp(pd.Period(max_ym,'M').start_time) - pd.DateOffset(months=2)
        ativo=ult_dt>=cutoff
        contracts.append({
            'tipo':'FUP','contato':contato,'produto':'Acompanhamento',
            'detalhe':'Ativo' if ativo else 'INATIVO','fa':fa,
            'val_mensal':val_med,'total_parc':999,'parc_atual':0,
            'data_ini':None,'monthly':monthly,'bonus_elegivel':False,'ativo':ativo
        })

    # ── W1 BUSINESS ──
    w1b=ent[ent['Empresa'].str.contains('Business',na=False)].copy()
    biz_val = params.get('business_valor', 0)
    if len(w1b)>0:
        monthly=w1b.groupby(w1b['YM'].astype(str))['VG'].sum().to_dict()
        contracts.append({
            'tipo':'W1 Business','contato':'Equipe','produto':'W1 Business',
            'detalhe':f'R${biz_val:,.0f}/mês proj.' if biz_val>0 else 'Variável',
            'fa':'FA IV','val_mensal':biz_val if biz_val>0 else 400.0,
            'total_parc':999,'parc_atual':0,'data_ini':None,'monthly':monthly,'bonus_elegivel':False
        })

    # ── W1 HOLDING ──
    w1h=ent[ent['Empresa'].str.contains('Holdings',na=False)].copy()
    h_parcelas = params.get('holding_parcelas', 0)
    h_valor = params.get('holding_valor', 0)
    h_pagas = params.get('holding_pagas', 0)
    if len(w1h)>0 or h_parcelas>0:
        monthly=w1h.groupby(w1h['YM'].astype(str))['VG'].sum().to_dict() if len(w1h)>0 else {}
        contracts.append({
            'tipo':'W1 Holding','contato':NM.split()[0],'produto':'W1 Holding',
            'detalhe':f'{h_parcelas} parc R${h_valor:,.0f}' if h_parcelas>0 else 'Variável',
            'fa':'FA IV','val_mensal':h_valor if h_valor>0 else 250.0,
            'total_parc':h_parcelas if h_parcelas>0 else 999,
            'parc_atual':h_pagas,'data_ini':None,'monthly':monthly,'bonus_elegivel':False,
            'holding_restam':max(0, h_parcelas - h_pagas)
        })

    # Sort
    type_order={'Consórcio':0,'MAG Vida':1,'Horizonte/Prev':2,'AP':3,'FUP':4,'W1 Business':5,'W1 Holding':6}
    contracts.sort(key=lambda x:(type_order.get(x['tipo'],99),x['contato'],x.get('produto','')))

    # ═══════════════════════════════════════════════════════════════════════════
    # PROJECTIONS (base)
    # ═══════════════════════════════════════════════════════════════════════════
    for ct in contracts:
        for pi,pym in enumerate(PROJ_YM):
            if ct['tipo']=='Consórcio':
                fp=ct['parc_atual']+pi+1
                if fp<=12: ct['monthly'][pym]=ct['val_mensal']
            elif ct['tipo']=='MAG Vida':
                fp=ct['parc_atual']+pi+1
                if fp<=12: ct['monthly'][pym]=ct['val_mensal']
                elif fp<=120: ct['monthly'][pym]=ct.get('val_y2',ct['val_mensal']*0.36)
            elif ct['tipo']=='Horizonte/Prev':
                ct['monthly'][pym]=ct['val_mensal']
            elif ct['tipo']=='AP':
                if 'À vista' in ct.get('detalhe',''):pass
                elif ct.get('ult_dt') and ct['ult_dt']<pd.Timestamp(pd.Period(max_ym,'M').start_time) - pd.DateOffset(months=2):pass
                else:
                    fp=ct['parc_atual']+pi+1
                    if fp<=12: ct['monthly'][pym]=ct['val_mensal']
            elif ct['tipo']=='FUP':
                if ct.get('ativo',False): ct['monthly'][pym]=ct['val_mensal']
            elif ct['tipo']=='W1 Business':
                ct['monthly'][pym]=ct['val_mensal']
            elif ct['tipo']=='W1 Holding':
                restam = ct.get('holding_restam', 0)
                if restam > 0 and pi < restam:
                    ct['monthly'][pym]=ct['val_mensal']

    # ═══════════════════════════════════════════════════════════════════════════
    # BONUS MATCHING
    # ═══════════════════════════════════════════════════════════════════════════
    first_b2000 = params.get('first_bonus_2000')
    first_b3000 = params.get('first_bonus_3000')

    for ct in contracts:
        ct['bonus_monthly']={}
        if not ct['bonus_elegivel']:
            continue

        contato_n=norm_name(ct['contato'])
        prod_n=norm_prod(ct['produto'])

        for base_ym in ALL_YM:
            base_val=ct['monthly'].get(base_ym,0)
            if base_val<=0:
                continue

            bonus_ym_idx=ALL_YM.index(base_ym)+1
            if bonus_ym_idx>=len(ALL_YM):
                continue
            bonus_ym=ALL_YM[bonus_ym_idx]

            fa_rate=FA_RATES.get(ct['fa'],0.175)
            exp_2000=round(base_val*(0.40-fa_rate)/fa_rate,2)
            exp_3000=round(base_val*(0.45-fa_rate)/fa_rate,2)

            is_hist=ALL_YM.index(bonus_ym)<N_HIST

            if is_hist:
                key=(bonus_ym, contato_n, prod_n)
                available=bonus_lookup.get(key,[])
                matched=None

                for entry in available:
                    if entry['used']: continue
                    if abs(entry['valor']-exp_2000)<max(1.0, exp_2000*0.12):
                        matched=entry; entry['used']=True
                        ct['bonus_monthly'][bonus_ym]={'recebido':entry['valor'],'esperado':exp_2000,'status':'OK','tipo':entry['tipo']}
                        break
                    elif abs(entry['valor']-exp_3000)<max(1.0, exp_3000*0.12):
                        matched=entry; entry['used']=True
                        ct['bonus_monthly'][bonus_ym]={'recebido':entry['valor'],'esperado':exp_3000,'status':'OK','tipo':entry['tipo']}
                        break

                if matched is None:
                    for entry in available:
                        if entry['used']:continue
                        matched=entry; entry['used']=True
                        exp=exp_2000 if abs(entry['valor']-exp_2000)<abs(entry['valor']-exp_3000) else exp_3000
                        ct['bonus_monthly'][bonus_ym]={'recebido':entry['valor'],'esperado':exp,'status':'DESVIO','tipo':entry['tipo']}
                        break

                if matched is None:
                    # Check if bonus should exist based on user-provided dates
                    bonus_expected = False
                    if first_b2000 and base_ym >= first_b2000:
                        bonus_expected = True
                    if first_b3000 and base_ym >= first_b3000:
                        bonus_expected = True
                    # Also check: only if within auditable range
                    if bonus_expected and base_ym <= LAST_BONUS_BASE:
                        ct['bonus_monthly'][bonus_ym]={'recebido':0,'esperado':exp_2000,'status':'FALTA','tipo':'-'}
            else:
                # Projected
                exp=exp_3000 if (first_b3000 and base_ym >= first_b3000) else exp_2000
                if first_b2000 and base_ym >= first_b2000:
                    ct['bonus_monthly'][bonus_ym]={'recebido':0,'esperado':exp,'status':'PROJ','tipo':'Proj'}
                elif first_b3000 and base_ym >= first_b3000:
                    ct['bonus_monthly'][bonus_ym]={'recebido':0,'esperado':exp,'status':'PROJ','tipo':'Proj'}

    # Stats
    n_ok=sum(1 for ct in contracts for b in ct['bonus_monthly'].values() if b.get('status')=='OK')
    n_dev=sum(1 for ct in contracts for b in ct['bonus_monthly'].values() if b.get('status')=='DESVIO')
    n_falta=sum(1 for ct in contracts for b in ct['bonus_monthly'].values() if b.get('status')=='FALTA')
    v_falta=sum(b['esperado'] for ct in contracts for b in ct['bonus_monthly'].values() if b.get('status')=='FALTA')

    # Bônus Alto Padrão
    ap_bonus_monthly={}
    if not params.get('exclude_ap_bonus', True):
        for _,r in b_ap.iterrows():
            ym=str(r['YM'])
            ap_bonus_monthly[ym]=ap_bonus_monthly.get(ym,0)+r['VG']

    # ═══════════════════════════════════════════════════════════════════════════
    # BUILD WORKBOOK
    # ═══════════════════════════════════════════════════════════════════════════
    wb=Workbook()

    # Current FA level
    current_fa = 'FA IV'
    if params.get('promos'):
        for lvl in ['FA IV','FA III','FA II','FA I']:
            if params['promos'].get(lvl):
                current_fa = lvl
                break

    bonus_pct = '45%' if first_b3000 else ('40%' if first_b2000 else '—')

    # ════════════════════ SHEET 1: FLUXO UNIFICADO ═══════════════════════════
    ws=wb.active; ws.title="Fluxo Completo"
    ws.sheet_view.showGridLines=False

    last_col=6+len(ALL_LBL)+3
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=last_col)
    c=ws['A1'];c.value=f"FLUXO UNIFICADO DE COMISSÕES + BÔNUS  |  {NM}  |  W1"
    c.font=Font(name=FN,bold=True,color="FFFFFF",size=13)
    c.fill=PatternFill("solid",fgColor=C_N)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=32

    # ── Row 2: legend with Holding/Business/Bonus info ──
    legend_parts = [
        f"Atualizado: {datetime.now().strftime('%d/%m/%Y')}",
        f"{current_fa}",
    ]
    if first_b3000:
        legend_parts.append(f"+3000 PPs = 45% (base a partir de {first_b3000})")
    if first_b2000:
        legend_parts.append(f"+2000 PPs = 40% (base a partir de {first_b2000})")
    if h_parcelas > 0:
        legend_parts.append(f"Holding: {h_parcelas}x R${h_valor:,.0f} ({h_pagas} pagas)")
    if biz_val > 0:
        legend_parts.append(f"Business: R${biz_val:,.0f}/mês")
    legend_parts.append("Azul=recebido | Verde=projetado | Vermelho=FALTA | Roxo=bônus | Bônus=M+1 da base")

    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=last_col)
    ws['A2'].value="  |  ".join(legend_parts)
    ws['A2'].font=Font(name=FN,size=7,color="FFFFFF",italic=True)
    ws['A2'].fill=PatternFill("solid",fgColor=C_B)
    ws['A2'].alignment=Alignment(horizontal="center",vertical="center")

    # ── Row 3: explanation of bonus calculation ──
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=last_col)
    bonus_explain = "COMO FUNCIONA O BÔNUS: contratos de Produtos Parceiros (Consórcio, MAG, Horizonte, Zurich) geram bônus no mês seguinte ao pagamento base."
    if first_b2000:
        bonus_explain += f"  Bônus 2000: fator = (40% - % FA) / % FA."
    if first_b3000:
        bonus_explain += f"  Bônus 3000: fator = (45% - % FA) / % FA."
    bonus_explain += "  Ex: Base R$100 como FA III (17.5%) → Bônus 3000 = R$100 × (0.45-0.175)/0.175 = R$157,14."
    ws['A3'].value=bonus_explain
    ws['A3'].font=Font(name=FN,size=6.5,color="333333",italic=True)
    ws['A3'].fill=PatternFill("solid",fgColor="E8EEF5")
    ws['A3'].alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    ws.row_dimensions[3].height=28

    # Headers row 4
    fixed_hdrs=['Tipo','Cliente','Produto','Detalhe','FA Congelado','Val. Base/mês']
    for i,h in enumerate(fixed_hdrs):
        hdr(ws,4,i+1,h,bg=C_N,sz=8)

    col_start=len(fixed_hdrs)+1
    for i,lbl in enumerate(ALL_LBL):
        bg=C_N if i<N_HIST else C_G
        hdr(ws,4,col_start+i,lbl,bg=bg,sz=7)

    hdr(ws,4,col_start+len(ALL_LBL),"Tot. Receb.",bg=C_N,sz=7)
    hdr(ws,4,col_start+len(ALL_LBL)+1,"Tot. Proj.",bg=C_G,sz=7)
    hdr(ws,4,col_start+len(ALL_LBL)+2,"TOTAL",bg=C_P,sz=7)

    # ── Write contract rows ──
    current_tipo=None
    row=5
    grand_base_hist=0; grand_base_proj=0; grand_bonus_hist=0; grand_bonus_proj=0

    for ci,ct in enumerate(contracts):
        if ct['tipo']!=current_tipo:
            current_tipo=ct['tipo']
            ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=col_start+len(ALL_LBL)+2)
            tipo_bg={'Consórcio':C_B,'MAG Vida':C_GL,'Horizonte/Prev':"3F7CAC",
                     'AP':"7B68EE",'FUP':"FF8C00",'W1 Business':"555555",'W1 Holding':"555555"}
            extra_label=" + Bônus" if ct['bonus_elegivel'] else ""
            c=ws.cell(row=row,column=1,value=f"▸  {current_tipo.upper()}{extra_label}")
            c.font=Font(name=FN,bold=True,color="FFFFFF",size=9)
            c.fill=PatternFill("solid",fgColor=tipo_bg.get(current_tipo,C_N))
            c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
            row+=1

        rbg=C_R1 if ci%2==0 else C_R2

        # ── LINE 1: BASE ──
        dc(ws,row,1,"Base",bg=rbg,al="center",bold=True,sz=7)
        dc(ws,row,2,ct['contato'][:28],bold=True,bg=rbg)
        dc(ws,row,3,ct['produto'][:22],bg=rbg)
        dc(ws,row,4,ct.get('detalhe',''),bg=rbg,al="center")
        dc(ws,row,5,ct['fa'],bg=rbg,al="center")
        dc(ws,row,6,ct['val_mensal'],fmt=FB,bg=rbg,al="right")

        total_base_r=0; total_base_p=0
        for mi,ym in enumerate(ALL_YM):
            col=col_start+mi
            val=ct['monthly'].get(ym)
            if val and val>0:
                if mi<N_HIST:
                    dc(ws,row,col,round(val,2),fmt=FB,bg="DCEEFB",al="right")
                    total_base_r+=val
                else:
                    dc(ws,row,col,round(val,2),fmt=FB,bg=C_LG,al="right")
                    total_base_p+=val
            else:
                if mi<N_HIST:
                    hist_months=[ym2 for ym2 in HIST_YM if ym2 in ct['monthly'] and ct['monthly'][ym2]>0]
                    if hist_months and min(hist_months)<=ym<=max(hist_months):
                        dc(ws,row,col,"FALTA",bg="FFCDD2",fg=C_R,al="center",bold=True,sz=7)
                    else:
                        dc(ws,row,col,None,bg=rbg)
                else:
                    dc(ws,row,col,None,bg=rbg)

        dc(ws,row,col_start+len(ALL_LBL),total_base_r,fmt=FB,bg="DCEEFB",al="right",bold=True)
        dc(ws,row,col_start+len(ALL_LBL)+1,total_base_p,fmt=FB,bg=C_LG,al="right",bold=True)
        dc(ws,row,col_start+len(ALL_LBL)+2,total_base_r+total_base_p,fmt=FB,bg="EDE7F6",al="right",bold=True)
        grand_base_hist+=total_base_r; grand_base_proj+=total_base_p
        row+=1

        # ── LINE 2: BÔNUS ──
        if ct['bonus_elegivel']:
            bonus_bg="F3E8FF"
            dc(ws,row,1,"Bônus",bg=bonus_bg,al="center",bold=True,fg=C_P,sz=7)
            dc(ws,row,2,ct['contato'][:28],bg=bonus_bg,fg="666666")
            dc(ws,row,3,ct['produto'][:22],bg=bonus_bg,fg="666666")
            dc(ws,row,4,"M+1",bg=bonus_bg,al="center",fg=C_P)
            dc(ws,row,5,ct['fa'],bg=bonus_bg,al="center",fg="666666")
            dc(ws,row,6,"",bg=bonus_bg)

            total_bonus_r=0; total_bonus_p=0
            for mi,ym in enumerate(ALL_YM):
                col=col_start+mi
                bdata=ct['bonus_monthly'].get(ym)
                if bdata:
                    if bdata['status']=='OK':
                        dc(ws,row,col,round(bdata['recebido'],2),fmt=FB,bg="E8DAEF",al="right")
                        total_bonus_r+=bdata['recebido']
                    elif bdata['status']=='DESVIO':
                        dc(ws,row,col,round(bdata['recebido'],2),fmt=FB,bg="FFF9C4",al="right",bold=True)
                        total_bonus_r+=bdata['recebido']
                    elif bdata['status']=='FALTA':
                        dc(ws,row,col,"FALTA",bg="FFCDD2",fg=C_R,al="center",bold=True,sz=7)
                    elif bdata['status']=='PROJ':
                        dc(ws,row,col,round(bdata['esperado'],2),fmt=FB,bg="D5F5E3",al="right")
                        total_bonus_p+=bdata['esperado']
                else:
                    dc(ws,row,col,None,bg=bonus_bg)

            dc(ws,row,col_start+len(ALL_LBL),total_bonus_r,fmt=FB,bg="E8DAEF",al="right",bold=True)
            dc(ws,row,col_start+len(ALL_LBL)+1,total_bonus_p,fmt=FB,bg="D5F5E3",al="right",bold=True)
            dc(ws,row,col_start+len(ALL_LBL)+2,total_bonus_r+total_bonus_p,fmt=FB,bg="EDE7F6",al="right",bold=True)
            grand_bonus_hist+=total_bonus_r; grand_bonus_proj+=total_bonus_p
            row+=1

    # ── BÔNUS ALTO PADRÃO ──
    if ap_bonus_monthly:
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=col_start+len(ALL_LBL)+2)
        c=ws.cell(row=row,column=1,value="▸  BÔNUS CLIENTE DE ALTO PADRÃO")
        c.font=Font(name=FN,bold=True,color="FFFFFF",size=9)
        c.fill=PatternFill("solid",fgColor=C_P)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        row+=1
        dc(ws,row,1,"Bônus AP",bg="F3E8FF",al="center",bold=True,fg=C_P,sz=7)
        dc(ws,row,2,"Clientes AP",bold=True,bg="F3E8FF")
        dc(ws,row,3,"Alto Padrão",bg="F3E8FF")
        for i in range(3,7): dc(ws,row,i,"",bg="F3E8FF")
        total_ap_r=0
        for mi,ym in enumerate(ALL_YM):
            col=col_start+mi
            v=ap_bonus_monthly.get(ym,0)
            if v>0 and mi<N_HIST:
                dc(ws,row,col,round(v,2),fmt=FB,bg="E8DAEF",al="right",bold=True)
                total_ap_r+=v
            else:
                dc(ws,row,col,None,bg="F3E8FF")
        dc(ws,row,col_start+len(ALL_LBL),total_ap_r,fmt=FB,bg="E8DAEF",al="right",bold=True)
        dc(ws,row,col_start+len(ALL_LBL)+1,0,fmt=FB,bg="D5F5E3",al="right")
        dc(ws,row,col_start+len(ALL_LBL)+2,total_ap_r,fmt=FB,bg="EDE7F6",al="right",bold=True)
        grand_bonus_hist+=total_ap_r
        row+=1

    # ── GRAND TOTAL ──
    row+=1
    dc(ws,row,1,"",bg=C_N);dc(ws,row,2,"TOTAL GERAL (Base + Bônus)",bold=True,bg=C_N,fg="FFFFFF")
    for ci2 in range(3,col_start): dc(ws,row,ci2,"",bg=C_N)

    for mi,ym in enumerate(ALL_YM):
        col=col_start+mi
        base_t=sum(ct['monthly'].get(ym,0) for ct in contracts)
        bonus_t=sum(ct['bonus_monthly'].get(ym,{}).get('recebido',0) for ct in contracts if ct['bonus_elegivel'])
        bonus_proj=sum(ct['bonus_monthly'].get(ym,{}).get('esperado',0) for ct in contracts if ct['bonus_elegivel'] and ct['bonus_monthly'].get(ym,{}).get('status')=='PROJ')
        ap_t=ap_bonus_monthly.get(ym,0) if mi<N_HIST else 0
        if mi<N_HIST:
            total=base_t+bonus_t+ap_t
            dc(ws,row,col,round(total,2),fmt=FB,bg=C_N,fg="FFFFFF",al="right",bold=True)
        else:
            total=base_t+bonus_proj
            dc(ws,row,col,round(total,2),fmt=FB,bg=C_G,fg="FFFFFF",al="right",bold=True)

    gt_r=grand_base_hist+grand_bonus_hist
    gt_p=grand_base_proj+grand_bonus_proj
    dc(ws,row,col_start+len(ALL_LBL),gt_r,fmt=FB,bg=C_N,fg="FFFFFF",al="right",bold=True)
    dc(ws,row,col_start+len(ALL_LBL)+1,gt_p,fmt=FB,bg=C_G,fg="FFFFFF",al="right",bold=True)
    dc(ws,row,col_start+len(ALL_LBL)+2,gt_r+gt_p,fmt=FB,bg=C_P,fg="FFFFFF",al="right",bold=True)

    borders(ws,4,row,1,col_start+len(ALL_LBL)+2)
    set_cols(ws,{'A':6,'B':22,'C':18,'D':12,'E':5,'F':10})
    for i in range(len(ALL_LBL)):
        ws.column_dimensions[get_column_letter(col_start+i)].width=9
    for extra in range(3):
        ws.column_dimensions[get_column_letter(col_start+len(ALL_LBL)+extra)].width=11
    ws.freeze_panes='G5'

    # ════════════════════ SHEET 2: BÔNUS FALTANTES ═══════════════════════════
    ws2=wb.create_sheet("Bônus Faltantes")
    ws2.sheet_view.showGridLines=False
    ws2.merge_cells('A1:H1')
    c=ws2['A1'];c.value="BÔNUS FALTANTES  •  Base paga mas bônus não recebido no mês seguinte  •  COBRAR W1"
    c.font=Font(name=FN,bold=True,color="FFFFFF",size=11)
    c.fill=PatternFill("solid",fgColor=C_R)
    c.alignment=Alignment(horizontal="center",vertical="center")

    h2=['Cliente','Produto','FA','Mês Base','Mês Bônus','Valor Base','Bônus Esperado','Status']
    for i,h in enumerate(h2):
        hdr(ws2,2,i+1,h,bg=C_N,sz=9)

    falta_list=[]
    for ct in contracts:
        if not ct['bonus_elegivel']:continue
        for ym,bdata in ct['bonus_monthly'].items():
            if bdata['status']=='FALTA':
                idx=ALL_YM.index(ym)-1 if ym in ALL_YM and ALL_YM.index(ym)>0 else -1
                base_ym=ALL_YM[idx] if idx>=0 else '?'
                base_val=ct['monthly'].get(base_ym,0)
                falta_list.append({
                    'contato':ct['contato'],'produto':ct['produto'],'fa':ct['fa'],
                    'base_ym':base_ym,'bonus_ym':ym,'base_val':base_val,
                    'esperado':bdata['esperado']
                })

    falta_list.sort(key=lambda x:(x['bonus_ym'],x['contato']))
    if falta_list:
        for ri,f in enumerate(falta_list):
            r2=ri+3; bg=C_R1 if ri%2==0 else C_R2
            dc(ws2,r2,1,f['contato'][:28],bold=True,bg=bg)
            dc(ws2,r2,2,f['produto'][:22],bg=bg)
            dc(ws2,r2,3,f['fa'],bg=bg,al="center")
            dc(ws2,r2,4,f['base_ym'],bg=bg,al="center")
            dc(ws2,r2,5,f['bonus_ym'],bg="FFCDD2",fg=C_R,al="center",bold=True)
            dc(ws2,r2,6,f['base_val'],fmt=FB,bg=bg,al="right")
            dc(ws2,r2,7,f['esperado'],fmt=FB,bg="FFCDD2",fg=C_R,al="right",bold=True)
            dc(ws2,r2,8,"COBRAR W1",bg="FFCDD2",fg=C_R,al="center",bold=True)
        borders(ws2,2,2+len(falta_list),1,8)
        sr=3+len(falta_list)+1
        total_f=sum(f['esperado'] for f in falta_list)
        dc(ws2,sr,1,f"TOTAL: {len(falta_list)} bônus faltando",bold=True,bg=C_R,fg="FFFFFF")
        dc(ws2,sr,7,total_f,fmt=FB,bold=True,bg=C_R,fg="FFFFFF",al="right")
    else:
        dc(ws2,3,1,"Nenhum bônus faltando!",bold=True,fg=C_G)
    set_cols(ws2,{'A':28,'B':22,'C':6,'D':9,'E':9,'F':11,'G':14,'H':12})

    # ════════════════════ SHEET 3: PARCELAS BASE FALTANTES ═══════════════════
    ws3=wb.create_sheet("Parcelas Base Faltantes")
    ws3.sheet_view.showGridLines=False
    ws3.merge_cells('A1:G1')
    c=ws3['A1'];c.value="PARCELAS BASE FALTANTES  •  Meses onde comissão base era esperada mas não veio"
    c.font=Font(name=FN,bold=True,color="FFFFFF",size=11)
    c.fill=PatternFill("solid",fgColor=C_R)
    c.alignment=Alignment(horizontal="center",vertical="center")

    h3=['Cliente','Produto','Tipo','Detalhe','Mês Faltante','Valor Esperado','Status']
    for i,h in enumerate(h3):
        hdr(ws3,2,i+1,h,bg=C_N,sz=9)

    missing_base=[]
    for ct in contracts:
        m=ct['monthly']
        hist_months=sorted([ym for ym in HIST_YM if ym in m and m[ym]>0])
        if len(hist_months)>=2:
            first_idx=HIST_YM.index(hist_months[0])
            last_idx=HIST_YM.index(hist_months[-1])
            for idx in range(first_idx,last_idx+1):
                ym=HIST_YM[idx]
                if ym not in m or m.get(ym,0)<=0:
                    missing_base.append({
                        'contato':ct['contato'],'produto':ct['produto'],
                        'tipo':ct['tipo'],'detalhe':ct.get('detalhe',''),
                        'mes':ALL_LBL[idx],'valor':ct['val_mensal']
                    })

    if missing_base:
        for ri,mb in enumerate(missing_base):
            r3=ri+3; bg=C_R1 if ri%2==0 else C_R2
            dc(ws3,r3,1,mb['contato'][:28],bold=True,bg=bg)
            dc(ws3,r3,2,mb['produto'][:22],bg=bg)
            dc(ws3,r3,3,mb['tipo'],bg=bg,al="center")
            dc(ws3,r3,4,mb['detalhe'],bg=bg,al="center")
            dc(ws3,r3,5,mb['mes'],bg="FFCDD2",fg=C_R,al="center",bold=True)
            dc(ws3,r3,6,mb['valor'],fmt=FB,bg=bg,al="right")
            dc(ws3,r3,7,"COBRAR W1",bg="FFCDD2",fg=C_R,al="center",bold=True)
        borders(ws3,2,2+len(missing_base),1,7)
        sr=3+len(missing_base)+1
        dc(ws3,sr,1,f"TOTAL: {len(missing_base)} parcelas faltando",bold=True,bg=C_R,fg="FFFFFF")
        dc(ws3,sr,6,sum(mb['valor'] for mb in missing_base),fmt=FB,bold=True,bg=C_R,fg="FFFFFF",al="right")
    set_cols(ws3,{'A':28,'B':22,'C':14,'D':14,'E':10,'F':14,'G':12})

    # ════════════════════ SHEET 4: RESUMO MENSAL ═════════════════════════════
    ws4=wb.create_sheet("Resumo Mensal")
    ws4.sheet_view.showGridLines=False
    ws4.merge_cells('A1:H1')
    c=ws4['A1'];c.value=f"RESUMO MENSAL  •  {NM}  •  Base + Bônus  •  Histórico + Projeção"
    c.font=Font(name=FN,bold=True,color="FFFFFF",size=12)
    c.fill=PatternFill("solid",fgColor=C_N)
    c.alignment=Alignment(horizontal="center",vertical="center")

    h4=['Mês','Base Receb.','Bônus Receb.','Total Receb.','Base Proj.','Bônus Proj.','Total Proj.','Bônus FALTA']
    for i,h in enumerate(h4):
        hdr(ws4,2,i+1,h,bg=C_B,sz=8)

    for mi,ym in enumerate(ALL_YM):
        r4=mi+3; bg=C_R1 if mi%2==0 else C_R2
        if mi>=N_HIST: bg=C_LG if mi%2==0 else "FFFFFF"
        dc(ws4,r4,1,ALL_LBL[mi],bold=True,bg=bg,al="center")
        base_t=sum(ct['monthly'].get(ym,0) for ct in contracts)
        bonus_r=sum(ct['bonus_monthly'].get(ym,{}).get('recebido',0) for ct in contracts if ct['bonus_elegivel'])
        bonus_r+=ap_bonus_monthly.get(ym,0) if mi<N_HIST else 0
        bonus_p=sum(ct['bonus_monthly'].get(ym,{}).get('esperado',0) for ct in contracts if ct['bonus_elegivel'] and ct['bonus_monthly'].get(ym,{}).get('status')=='PROJ')
        bonus_f=sum(ct['bonus_monthly'].get(ym,{}).get('esperado',0) for ct in contracts if ct['bonus_elegivel'] and ct['bonus_monthly'].get(ym,{}).get('status')=='FALTA')

        if mi<N_HIST:
            dc(ws4,r4,2,base_t,fmt=FB,bg=bg,al="right")
            dc(ws4,r4,3,bonus_r if bonus_r>0 else None,fmt=FB,bg=bg,al="right")
            dc(ws4,r4,4,base_t+bonus_r,fmt=FB,bg=bg,al="right",bold=True)
            dc(ws4,r4,5,None,bg=bg); dc(ws4,r4,6,None,bg=bg); dc(ws4,r4,7,None,bg=bg)
        else:
            dc(ws4,r4,2,None,bg=bg); dc(ws4,r4,3,None,bg=bg); dc(ws4,r4,4,None,bg=bg)
            dc(ws4,r4,5,base_t,fmt=FB,bg=bg,al="right")
            dc(ws4,r4,6,bonus_p if bonus_p>0 else None,fmt=FB,bg=bg,al="right")
            dc(ws4,r4,7,base_t+bonus_p,fmt=FB,bg=bg,al="right",bold=True)
        dc(ws4,r4,8,bonus_f if bonus_f>0 else None,fmt=FB,bg="FFCDD2" if bonus_f>0 else bg,fg=C_R if bonus_f>0 else "000000",al="right",bold=bonus_f>0)

    borders(ws4,2,2+len(ALL_YM),1,8)
    set_cols(ws4,{'A':8,'B':13,'C':13,'D':13,'E':13,'F':13,'G':13,'H':13})

    # ════════════════════ SHEET 5: COMPARAÇÃO COM EXTRATO ════════════════════
    ws5=wb.create_sheet("Comparação Extrato")
    ws5.sheet_view.showGridLines=False
    ws5.merge_cells('A1:G1')
    c=ws5['A1'];c.value="COMPARAÇÃO: PLANILHA vs EXTRATO  •  Validação por mês  •  Verificar cargos e valores"
    c.font=Font(name=FN,bold=True,color="FFFFFF",size=11)
    c.fill=PatternFill("solid",fgColor=C_N)
    c.alignment=Alignment(horizontal="center",vertical="center")

    # Explanation row
    ws5.merge_cells('A2:G2')
    ws5['A2'].value="Compare o valor total da planilha (base+bônus) com o que aparece no extrato CSV. Diferenças podem indicar cargo errado ou valor faltante."
    ws5['A2'].font=Font(name=FN,size=7,color="333333",italic=True)
    ws5['A2'].fill=PatternFill("solid",fgColor="E8EEF5")
    ws5['A2'].alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)

    for i,h in enumerate(['Mês','Extrato Total','Planilha Total','Diferença','% Diff','FA no Mês','Status']):
        hdr(ws5,3,i+1,h,bg=C_B,sz=8)

    for mi,ym in enumerate(HIST_YM):
        r5=mi+4; bg=C_R1 if mi%2==0 else C_R2
        dc(ws5,r5,1,ALL_LBL[mi],bold=True,bg=bg,al="center")

        real=ent[ent['YM'].astype(str)==ym]['VG'].sum()
        base_t=sum(ct['monthly'].get(ym,0) for ct in contracts)
        bonus_r=sum(ct['bonus_monthly'].get(ym,{}).get('recebido',0) for ct in contracts if ct['bonus_elegivel'])
        bonus_r+=ap_bonus_monthly.get(ym,0)
        planilha=base_t+bonus_r
        diff=real-planilha
        pct=abs(diff/real*100) if real>0 else 0

        dc(ws5,r5,2,round(real,2),fmt=FB,bg=bg,al="right")
        dc(ws5,r5,3,round(planilha,2),fmt=FB,bg=bg,al="right")
        diff_bg="FFCDD2" if abs(diff)>50 else bg
        dc(ws5,r5,4,round(diff,2),fmt=FB,bg=diff_bg,al="right",bold=abs(diff)>50)
        dc(ws5,r5,5,f"{pct:.1f}%",bg=diff_bg,al="center")

        # Show which FA positions appear in this month
        fa_counts=ent[ent['YM'].astype(str)==ym]['FA'].value_counts().to_dict()
        fa_str=", ".join([f"{k}({v})" for k,v in sorted(fa_counts.items()) if pd.notna(k)])
        dc(ws5,r5,6,fa_str,bg=bg,al="left",wrap=True)

        status="OK" if abs(diff)<50 else ("VERIFICAR" if abs(diff)<200 else "DIVERGÊNCIA")
        status_bg=bg if status=="OK" else ("FFF9C4" if status=="VERIFICAR" else "FFCDD2")
        dc(ws5,r5,7,status,bg=status_bg,al="center",bold=status!="OK")

    borders(ws5,3,3+len(HIST_YM),1,7)
    set_cols(ws5,{'A':8,'B':13,'C':13,'D':11,'E':7,'F':30,'G':12})

    # ════════════════════ SHEET 6: HISTÓRICO COMPLETO ════════════════════════
    ws6=wb.create_sheet("Histórico Completo")
    ws6.sheet_view.showGridLines=False
    ws6.merge_cells('A1:I1')
    c=ws6['A1'];c.value=f"HISTÓRICO COMPLETO  •  Todos os lançamentos  •  {ALL_LBL[0]} – {ALL_LBL[N_HIST-1]}"
    c.font=Font(name=FN,bold=True,color="FFFFFF",size=11)
    c.fill=PatternFill("solid",fgColor=C_N)
    c.alignment=Alignment(horizontal="center",vertical="center")

    for i,h in enumerate(['Data','Mês','Contato','Categoria','Produto','Parcela','FA','Obs','Valor']):
        hdr(ws6,2,i+1,h,bg=C_B,sz=8)

    ml=dict(zip(HIST_YM,ALL_LBL[:N_HIST]))
    for ri,(_,r) in enumerate(ent.sort_values('DT',ascending=False).iterrows()):
        r6=ri+3;bg=C_R1 if ri%2==0 else C_R2
        dc(ws6,r6,1,r['DT'].date() if pd.notna(r['DT']) else None,fmt='DD/MM/YYYY',bg=bg,al="center")
        dc(ws6,r6,2,ml.get(str(r['YM']),''),bg=bg,al="center")
        dc(ws6,r6,3,str(r['Contato'])[:30],bg=bg)
        dc(ws6,r6,4,str(r['Categoria'])[:22],bg=bg)
        dc(ws6,r6,5,str(r['Produto / Serviço'])[:22],bg=bg)
        dc(ws6,r6,6,str(r['Parcela']),bg=bg,al="center")
        dc(ws6,r6,7,str(r['FA']) if pd.notna(r['FA']) else '-',bg=bg,al="center")
        obs=str(r.get('Observações',''))[:45] if pd.notna(r.get('Observações')) else ''
        dc(ws6,r6,8,obs,bg=bg,wrap=True)
        dc(ws6,r6,9,r['VG'],fmt=FB,bg=bg,al="right")
    set_cols(ws6,{'A':12,'B':7,'C':26,'D':18,'E':20,'F':6,'G':6,'H':28,'I':11})

    # ── SAVE TO BYTES ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Monthly comparison for summary
    comparisons = []
    for mi,ym in enumerate(HIST_YM[-3:]):
        real=ent[ent['YM'].astype(str)==ym]['VG'].sum()
        base_t=sum(ct['monthly'].get(ym,0) for ct in contracts)
        bonus_r=sum(ct['bonus_monthly'].get(ym,{}).get('recebido',0) for ct in contracts if ct['bonus_elegivel'])
        bonus_r+=ap_bonus_monthly.get(ym,0)
        comparisons.append({'mes': ALL_LBL[ALL_YM.index(ym)], 'extrato': real, 'planilha': base_t+bonus_r})

    summary = {
        'nome': NM,
        'contratos': len(contracts),
        'linhas': sum(2 if ct['bonus_elegivel'] else 1 for ct in contracts),
        'base_hist': grand_base_hist,
        'bonus_hist': grand_bonus_hist,
        'total_hist': grand_base_hist + grand_bonus_hist,
        'base_proj': grand_base_proj,
        'bonus_proj': grand_bonus_proj,
        'total_proj': grand_base_proj + grand_bonus_proj,
        'bonus_ok': n_ok,
        'bonus_desvio': n_dev,
        'bonus_falta': n_falta,
        'bonus_falta_valor': v_falta,
        'parcelas_faltantes': len(missing_base),
        'comparisons': comparisons,
        'meses': f"{ALL_LBL[0]} – {ALL_LBL[-1]}",
        'hist_range': f"{ALL_LBL[0]} – {ALL_LBL[N_HIST-1]}",
    }

    return buf.getvalue(), summary


# ═══════════════════════════════════════════════════════════════════════════════
# STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="W1 Comissões – Controle Unificado",
    page_icon="📊",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1B3A5C, #2E6DA4);
        padding: 20px; border-radius: 10px; margin-bottom: 20px;
        color: white; text-align: center;
    }
    .main-header h1 { color: white; margin: 0; font-size: 1.8em; }
    .main-header p { color: #B8D4E8; margin: 5px 0 0 0; font-size: 0.9em; }
    .stat-card {
        background: #f8f9fa; padding: 15px; border-radius: 8px;
        border-left: 4px solid #2E6DA4; margin: 5px 0;
    }
    .stat-card h3 { margin: 0; color: #1B3A5C; font-size: 0.85em; }
    .stat-card p { margin: 5px 0 0 0; font-size: 1.3em; font-weight: bold; color: #2E6DA4; }
    .warning-card {
        background: #fff3cd; padding: 15px; border-radius: 8px;
        border-left: 4px solid #C0392B; margin: 5px 0;
    }
    .warning-card h3 { margin: 0; color: #C0392B; font-size: 0.85em; }
    .warning-card p { margin: 5px 0 0 0; font-size: 1.3em; font-weight: bold; color: #C0392B; }
    .info-box {
        background: #e8eef5; padding: 15px; border-radius: 8px;
        border-left: 4px solid #6B3FA0; margin: 10px 0;
    }
    .stExpander { border: 1px solid #ddd; border-radius: 8px; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>W1 Consultoria — Controle Unificado de Comissões</h1>
    <p>Faça upload do seu extrato CSV e configure os dados abaixo para gerar a planilha completa com base + bônus mês a mês</p>
</div>
""", unsafe_allow_html=True)

# ── SIDEBAR: Instructions ──
with st.sidebar:
    st.markdown("### Como usar")
    st.markdown("""
    1. **Upload** o extrato CSV do sistema W1
    2. **Preencha** seu nome exatamente como aparece no extrato
    3. **Configure** os dados de promoção e bônus
    4. Clique em **Gerar Planilha**
    5. **Baixe** o arquivo .xlsx
    """)

    st.markdown("---")
    st.markdown("### Entenda o Bônus")
    st.markdown("""
    **Bônus 2000 PPs** (comissão total = 40%):
    > Fator = (0.40 - % FA) / % FA

    **Bônus 3000 PPs** (comissão total = 45%):
    > Fator = (0.45 - % FA) / % FA

    **São mutuamente exclusivos**: cada entrada base gera OU Bônus 2000 OU Bônus 3000, nunca os dois.

    **Timing M+1**: O bônus referente à parcela paga no mês X aparece no mês X+1.
    Exemplo: parcela base paga em Set/25 → bônus cai em Out/25.

    Mesmo na última parcela (ex: parcela 12), o bônus vem no mês seguinte.
    """)

    st.markdown("---")
    st.markdown("### Tabela de % por Cargo")
    st.markdown("""
    | Cargo | % Base |
    |-------|--------|
    | FA I  | 7,5%   |
    | FA II | 15,0%  |
    | FA III| 17,5%  |
    | FA IV | 20,0%  |

    O % é **congelado** na data do contrato.
    """)

    st.markdown("---")
    st.markdown("### Produtos Elegíveis a Bônus")
    st.markdown("Consórcio, MAG Vida, Horizonte/Prev, Zurich")
    st.markdown("**NÃO** têm bônus: AP (Assessoria), FUP (Acompanhamento)")

# ── MAIN FORM ──
uploaded = st.file_uploader("📁 Upload do extrato CSV (semicolon-separated, latin-1)", type=['csv'])

if uploaded:
    csv_bytes = uploaded.getvalue()

    # Try to detect name from CSV (W1 exports are cp1252/latin-1)
    for enc in ['utf-8-sig', 'cp1252', 'latin-1']:
        try:
            temp_df = pd.read_csv(io.BytesIO(csv_bytes), sep=';', encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        temp_df = pd.read_csv(io.BytesIO(csv_bytes), sep=';', encoding='latin-1', encoding_errors='replace')

    # Find unique names across FA columns
    possible_names = set()
    for col in ['Nome FA I','Nome FA II','Nome FA III','Nome FA IV']:
        if col in temp_df.columns:
            vals = temp_df[col].dropna().unique()
            for v in vals:
                if str(v).strip() and str(v).strip() != '-':
                    possible_names.add(str(v).strip())

    st.markdown("---")
    st.markdown("### Dados do Consultor")

    col1, col2 = st.columns(2)

    with col1:
        if possible_names:
            nome = st.selectbox("Nome (como aparece no extrato)", sorted(possible_names),
                               help="Selecione o nome exatamente como aparece nas colunas 'Nome FA' do extrato")
        else:
            nome = st.text_input("Nome completo (como aparece no extrato)",
                                help="Digite o nome exatamente como aparece nas colunas 'Nome FA' do extrato")

    with col2:
        current_fa = st.selectbox("Cargo atual", ['FA IV','FA III','FA II','FA I'], index=0)

    st.markdown("---")
    st.markdown("### Datas de Promoção")
    st.markdown("*Informe o mês em que começou cada cargo (formato AAAA-MM). Deixe vazio se não se aplica.*")

    pc1, pc2, pc3 = st.columns(3)
    with pc1:
        promo_fa2 = st.text_input("Início FA II (AAAA-MM)", placeholder="2025-04",
                                   help="Mês em que foi promovido a FA II")
    with pc2:
        promo_fa3 = st.text_input("Início FA III (AAAA-MM)", placeholder="2025-08",
                                   help="Mês em que foi promovido a FA III")
    with pc3:
        promo_fa4 = st.text_input("Início FA IV (AAAA-MM)", placeholder="2026-02",
                                   help="Mês em que foi promovido a FA IV")

    st.markdown("---")
    st.markdown("### Bônus PPs")
    st.markdown("""
    *Informe o primeiro mês **base** cujos contratos geraram bônus. Lembre-se: o bônus é pago no mês seguinte (M+1).*

    *Exemplo: se fechou contratos em Ago/25, recebeu primeira parcela em Set/25, e o bônus apareceu em Out/25 → o mês base é **2025-09** (Set/25).*
    """)

    bc1, bc2 = st.columns(2)
    with bc1:
        first_b2000 = st.text_input("Primeiro mês base com Bônus 2000 PPs (AAAA-MM)",
                                     placeholder="2025-10",
                                     help="Primeiro mês BASE que gerou bônus de 2000 PPs (o bônus será pago no mês seguinte)")
    with bc2:
        first_b3000 = st.text_input("Primeiro mês base com Bônus 3000 PPs (AAAA-MM)",
                                     placeholder="2025-09",
                                     help="Primeiro mês BASE que gerou bônus de 3000 PPs (o bônus será pago no mês seguinte)")

    st.markdown("---")
    st.markdown("### W1 Holding & Business")

    hc1, hc2, hc3 = st.columns(3)
    with hc1:
        holding_parcelas = st.number_input("Holding: total de parcelas", min_value=0, value=0,
                                            help="Número total de parcelas da comissão de Holding (0 = sem Holding)")
    with hc2:
        holding_valor = st.number_input("Holding: valor por parcela (R$)", min_value=0.0, value=0.0, step=100.0,
                                         help="Valor de cada parcela de Holding")
    with hc3:
        holding_pagas = st.number_input("Holding: parcelas já pagas", min_value=0, value=0,
                                         help="Quantas parcelas de Holding já foram recebidas")

    bc1b, bc2b = st.columns(2)
    with bc1b:
        business_valor = st.number_input("Business: valor mensal projetado (R$)", min_value=0.0, value=0.0, step=100.0,
                                          help="Valor mensal de W1 Business (0 = sem Business)")

    st.markdown("---")

    # Advanced options
    with st.expander("Opções avançadas"):
        exclude_ap = st.checkbox("Excluir Bônus Cliente de Alto Padrão da planilha", value=True,
                                  help="Marque se o consultor não tem bônus de alto padrão")

    # ── GENERATE BUTTON ──
    if st.button("🚀 Gerar Planilha Unificada", type="primary", use_container_width=True):
        if not nome:
            st.error("Preencha o nome do consultor!")
        else:
            with st.spinner("Processando extrato e gerando planilha..."):
                params = {
                    'nome': nome,
                    'promos': {
                        'FA II': promo_fa2 if promo_fa2 else None,
                        'FA III': promo_fa3 if promo_fa3 else None,
                        'FA IV': promo_fa4 if promo_fa4 else None,
                    },
                    'first_bonus_2000': first_b2000 if first_b2000 else None,
                    'first_bonus_3000': first_b3000 if first_b3000 else None,
                    'holding_parcelas': int(holding_parcelas),
                    'holding_valor': float(holding_valor),
                    'holding_pagas': int(holding_pagas),
                    'business_valor': float(business_valor),
                    'exclude_ap_bonus': exclude_ap,
                }

                xlsx_bytes, summary = build_unified(csv_bytes, params)

                if xlsx_bytes is None:
                    st.error(summary.get('error', 'Erro ao processar o extrato.'))
                else:
                    st.success("Planilha gerada com sucesso!")
                    log_usage(summary['nome'], params.get('promos', ['FA I'])[0], summary['contratos'], summary['total_hist'], summary['total_proj'])

                    # ── RESULTS DASHBOARD ──
                    st.markdown(f"### Resultados: {summary['nome']}")
                    st.markdown(f"*Período: {summary['meses']}  |  Histórico: {summary['hist_range']}*")

                    # Stats cards
                    r1c1, r1c2, r1c3, r1c4 = st.columns(4)
                    with r1c1:
                        st.markdown(f"""<div class="stat-card"><h3>Contratos</h3><p>{summary['contratos']}</p></div>""", unsafe_allow_html=True)
                    with r1c2:
                        st.markdown(f"""<div class="stat-card"><h3>Total Histórico</h3><p>R$ {summary['total_hist']:,.2f}</p></div>""", unsafe_allow_html=True)
                    with r1c3:
                        st.markdown(f"""<div class="stat-card"><h3>Total Projetado</h3><p>R$ {summary['total_proj']:,.2f}</p></div>""", unsafe_allow_html=True)
                    with r1c4:
                        st.markdown(f"""<div class="stat-card"><h3>Linhas no Fluxo</h3><p>{summary['linhas']}</p></div>""", unsafe_allow_html=True)

                    r2c1, r2c2, r2c3 = st.columns(3)
                    with r2c1:
                        st.markdown(f"""<div class="stat-card"><h3>Bônus OK</h3><p>{summary['bonus_ok']}</p></div>""", unsafe_allow_html=True)
                    with r2c2:
                        st.markdown(f"""<div class="stat-card"><h3>Bônus com Desvio</h3><p>{summary['bonus_desvio']}</p></div>""", unsafe_allow_html=True)
                    with r2c3:
                        st.markdown(f"""<div class="warning-card"><h3>Bônus Faltantes</h3><p>{summary['bonus_falta']} (R$ {summary['bonus_falta_valor']:,.2f})</p></div>""", unsafe_allow_html=True)

                    # Parcelas faltantes
                    if summary['parcelas_faltantes'] > 0:
                        st.markdown(f"""<div class="warning-card"><h3>Parcelas Base Faltantes</h3><p>{summary['parcelas_faltantes']} parcelas não recebidas dentro do período esperado</p></div>""", unsafe_allow_html=True)

                    # Comparison table
                    if summary.get('comparisons'):
                        st.markdown("#### Validação: Planilha vs Extrato (últimos 3 meses)")
                        comp_df = pd.DataFrame(summary['comparisons'])
                        comp_df.columns = ['Mês', 'Extrato (R$)', 'Planilha (R$)']
                        comp_df['Diferença (R$)'] = comp_df['Extrato (R$)'] - comp_df['Planilha (R$)']
                        st.dataframe(comp_df.style.format({
                            'Extrato (R$)': 'R$ {:,.2f}',
                            'Planilha (R$)': 'R$ {:,.2f}',
                            'Diferença (R$)': 'R$ {:,.2f}'
                        }), use_container_width=True, hide_index=True)

                    st.markdown("---")

                    # Download button
                    fname = f"Controle_Comissoes_W1_{nome.replace(' ','_')}_UNIFICADO.xlsx"
                    st.download_button(
                        label="📥 Baixar Planilha (.xlsx)",
                        data=xlsx_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
                        type="primary",
                        use_container_width=True
                    )

                    st.markdown("""
                    <div class="info-box">
                        <strong>Abas da planilha:</strong><br>
                        1. <strong>Fluxo Completo</strong> – Cada contrato com linha Base + Bônus (M+1), mês a mês<br>
                        2. <strong>Bônus Faltantes</strong> – Lista de bônus esperados que não foram pagos → COBRAR W1<br>
                        3. <strong>Parcelas Base Faltantes</strong> – Parcelas de comissão base que não apareceram no extrato<br>
                        4. <strong>Resumo Mensal</strong> – Totais mensais de base + bônus + projeção<br>
                        5. <strong>Comparação Extrato</strong> – Validação automática: planilha vs extrato CSV (cargo e valor)<br>
                        6. <strong>Histórico Completo</strong> – Todos os lançamentos do extrato
                    </div>
                    """, unsafe_allow_html=True)

else:
    # Show welcome info when no file uploaded
    st.markdown("""
    <div class="info-box">
        <strong>Para começar:</strong> faça upload do extrato CSV de comissões extraído do sistema W1.<br><br>
        O arquivo é separado por ponto-e-vírgula (;) e possui colunas como: Data da Movimentação, Contato,
        Produto / Serviço, Parcela, Nome FA I/II/III/IV, Valor FA I/II/III/IV, etc.<br><br>
        <strong>O que a planilha gerada inclui:</strong><br>
        • Fluxo mês a mês de cada contrato (Consórcio, MAG, Horizonte, AP, FUP, Holding, Business)<br>
        • Linha de bônus pareada abaixo de cada contrato elegível (M+1 da base)<br>
        • Detecção de parcelas e bônus faltantes (para cobrar da W1)<br>
        • Projeção dos próximos 9 meses<br>
        • Validação automática contra o extrato CSV original
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### Exemplo de Fluxo")
    st.markdown("""
    ```
    ┌─────────┬──────────────────┬─────────────┬──────┬────────┬────────┬────────┐
    │  Tipo   │  Cliente         │  Produto    │  FA  │ Set/25 │ Out/25 │ Nov/25 │
    ├─────────┼──────────────────┼─────────────┼──────┼────────┼────────┼────────┤
    │  Base   │  Carolini Neri   │  Klubi Imóv │ FIII │ R$ 117 │ R$ 117 │ R$ 117 │
    │  Bônus  │  Carolini Neri   │  Klubi Imóv │ FIII │        │ R$ 183 │ R$ 150 │
    │  Base   │  João Silva      │  MAG Vida   │ FIV  │ R$ 200 │ R$ 200 │ R$ 200 │
    │  Bônus  │  João Silva      │  MAG Vida   │ FIV  │        │ R$ 250 │ R$ 250 │
    └─────────┴──────────────────┴─────────────┴──────┴────────┴────────┴────────┘

    Bônus de Out/25 refere-se à base de Set/25 (M+1)
    Bônus 3000 (FA III): R$117 × (0.45-0.175)/0.175 = R$183
    Bônus 2000 (FA III): R$117 × (0.40-0.175)/0.175 = R$150
    ```
    """)
